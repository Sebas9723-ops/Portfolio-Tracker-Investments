"""
POST /api/agents/broker-reconcile
Broker Reconciliation Agent (XTB):
  Strategy: wipe-and-reload
  1. Parse the XTB xlsx (Cash Operations)
  2. Delete all XTB-sourced transactions for this user
  3. Re-insert every trade from the file (BUY + SELL)
  4. FIFO-reconcile positions from the full transaction history
  5. Upsert positions table to match
  6. Validate with Groq
  One upload = correct state. No duplicate logic needed.
"""
from __future__ import annotations

import io
import logging
import re

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel

from app.auth.dependencies import get_user_id
from app.db.supabase_client import get_admin_client

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/agents", tags=["agents"])


# ── Ticker helpers ────────────────────────────────────────────────────────────

def _xtb_to_app_ticker(xtb: str) -> str:
    xtb = xtb.strip()
    if xtb.endswith(".US"):
        return xtb[:-3]            # VOO.US  → VOO
    if xtb.endswith(".UK"):
        return xtb[:-3] + ".L"    # EIMI.UK → EIMI.L
    return xtb                     # QDVE.DE → QDVE.DE


def _ticker_exchange_info(ticker: str) -> dict:
    if ticker.endswith(".DE"):
        return {"currency": "EUR", "market": "XETRA"}
    if ticker.endswith(".L"):
        return {"currency": "GBP", "market": "LSE"}
    return {"currency": "USD", "market": "US"}


def _ticker_aliases(ticker: str) -> list[str]:
    """All formats a ticker might be stored under in positions."""
    aliases = [ticker]
    if ticker.endswith(".L"):
        aliases.append(ticker[:-2] + ".UK")
    elif ticker.endswith(".UK"):
        aliases.append(ticker[:-3] + ".L")
    elif "." not in ticker:
        aliases.append(ticker + ".US")
    elif ticker.endswith(".US"):
        aliases.append(ticker[:-3])
    return aliases


# ── XTB parser ────────────────────────────────────────────────────────────────

def _parse_xtb_comment(comment: str) -> tuple[float, float] | None:
    """Extract (quantity, price) from XTB comment field."""
    m = re.search(
        r'(?:OPEN|CLOSE)\s+(?:BUY|SELL)\s+([\d.]+)(?:/[\d.]+)?\s*@\s*([\d.]+)',
        comment or "",
    )
    if not m:
        return None
    return float(m.group(1)), float(m.group(2))


def _parse_xtb_xlsx(content: bytes) -> tuple[list[dict], float]:
    """
    Parse XTB Cash Operations xlsx.
    Returns (trades, deposits_total_usd).
    trades contain BUY and SELL actions.
    """
    import openpyxl
    from datetime import datetime

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Find header row (contains "Type")
    header_row_idx = None
    col_map: dict[str, int] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Type" in cells:
            header_row_idx = i
            col_map = {name: idx for idx, name in enumerate(cells) if name}
            break
        if i > 30:
            break

    if header_row_idx is None:
        raise ValueError(
            "Header row not found. Export via XTB → Mi Cuenta → Historial "
            "→ Cash Operations → Exportar Excel."
        )

    def col(name: str, default: int) -> int:
        return col_map.get(name, default)

    idx_type    = col("Type",    0)
    idx_symbol  = col("Symbol",  1)
    idx_time    = col("Time",    3)
    idx_amount  = col("Amount",  4)
    idx_comment = col("Comment", 6)

    trades: list[dict] = []
    deposits_total = 0.0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or row[idx_type] is None:
            continue

        tx_type    = str(row[idx_type]).strip()
        ticker_raw = str(row[idx_symbol] or "").strip() if idx_symbol < len(row) else ""
        time_val   = row[idx_time]   if idx_time   < len(row) else None
        amount     = row[idx_amount] if idx_amount < len(row) else None
        comment    = str(row[idx_comment] or "").strip() if idx_comment < len(row) else ""

        # Deposits (case-insensitive)
        if tx_type.lower() == "deposit":
            try:
                deposits_total += float(amount or 0)
            except Exception:
                pass
            continue

        # Only process stock buys and sells
        if tx_type not in ("Stock purchase", "Stock sale", "Stock sell"):
            continue
        if not ticker_raw:
            continue

        parsed = _parse_xtb_comment(comment)
        if parsed is None:
            log.warning("Could not parse XTB comment: %s", comment[:80])
            continue
        qty, price = parsed

        if isinstance(time_val, datetime):
            tx_date = time_val.date().isoformat()
        else:
            tx_date = str(time_val)[:10]

        trades.append({
            "ticker_xtb": ticker_raw,
            "ticker":     _xtb_to_app_ticker(ticker_raw),
            "action":     "BUY" if tx_type == "Stock purchase" else "SELL",
            "quantity":   qty,
            "price_native": price,
            "date":       tx_date,
            "comment_raw": comment,
        })

    return trades, deposits_total


# ── Position reconciliation (FIFO) ────────────────────────────────────────────

def _reconcile_positions(user_id: str, db) -> dict[str, dict]:
    """
    Read ALL transactions for this user (ordered by date) and compute
    net shares + avg cost via FIFO lot accounting — matching XTB's display.
    Returns {ticker: {shares, avg_cost_native, currency, market}}.
    """
    tx_res = (
        db.table("transactions")
        .select("ticker,date,action,quantity,price_native")
        .eq("user_id", user_id)
        .order("date")
        .execute()
    )
    all_txs = tx_res.data or []

    # lots[ticker] = [[remaining_qty, price], ...]
    lots: dict[str, list[list[float]]] = {}

    for tx in all_txs:
        if tx.get("action") not in ("BUY", "SELL"):
            continue
        t   = tx["ticker"]
        qty = float(tx.get("quantity", 0))
        prc = float(tx.get("price_native", 0))

        if tx["action"] == "BUY":
            lots.setdefault(t, []).append([qty, prc])
        else:  # SELL — consume oldest lots first
            remaining = qty
            for lot in lots.get(t, []):
                if remaining <= 0:
                    break
                consumed   = min(lot[0], remaining)
                lot[0]    -= consumed
                remaining -= consumed

    result: dict[str, dict] = {}
    for ticker, lot_list in lots.items():
        open_lots  = [(q, p) for q, p in lot_list if q > 1e-8]
        total_qty  = sum(q for q, p in open_lots)
        total_cost = sum(q * p for q, p in open_lots)
        avg_cost   = total_cost / total_qty if total_qty > 0 else 0
        ex = _ticker_exchange_info(ticker)
        result[ticker] = {
            "shares":          round(total_qty, 6),
            "avg_cost_native": round(avg_cost, 6),
            "currency":        ex["currency"],
            "market":          ex["market"],
        }

    return result


# ── Name enrichment ──────────────────────────────────────────────────────────

def _enrich_names(tickers: list[str], db, user_id: str) -> None:
    """
    For every ticker in the list that has no name in positions,
    fetch the human-readable name from yfinance and write it back to DB.
    Uses pair fallback: if primary yf symbol returns nothing, tries alternatives.
    """
    import yfinance as yf
    from app.services.exchange_classifier import yf_ticker, get_ticker_pairs

    if not tickers:
        return

    # Load existing names to avoid unnecessary API calls
    pos_res = (
        db.table("positions")
        .select("ticker,name")
        .eq("user_id", user_id)
        .in_("ticker", tickers)
        .execute()
    )
    name_map = {p["ticker"]: p.get("name") for p in (pos_res.data or [])}

    for ticker in tickers:
        if name_map.get(ticker):  # already has a name
            continue

        # Build ordered list of yfinance symbols to try
        yf_symbols_to_try = [yf_ticker(ticker)] + [
            yf_ticker(p) for p in get_ticker_pairs(ticker)
        ]

        name = None
        for yf_sym in yf_symbols_to_try:
            try:
                info = yf.Ticker(yf_sym).info
                name = info.get("shortName") or info.get("longName")
                if name:
                    break
            except Exception:
                continue

        if name:
            try:
                db.table("positions") \
                  .update({"name": name}) \
                  .eq("user_id", user_id) \
                  .eq("ticker", ticker) \
                  .execute()
                log.info("Enriched name for %s: %s", ticker, name)
            except Exception as exc:
                log.warning("Could not write name for %s: %s", ticker, exc)


# ── Groq validation ───────────────────────────────────────────────────────────

def _groq_validate(trades: list[dict], reconciled: dict[str, dict], deposits_usd: float) -> str | None:
    try:
        from app.services.agent_pipeline import _call_groq

        ticker_lines = "\n".join(
            f"  • {t}: {v['shares']:.4f} shares @ avg {v['currency']} {v['avg_cost_native']:.4f}"
            for t, v in reconciled.items()
            if v["shares"] > 0
        )
        trade_lines = "\n".join(
            f"  • {tx['date']} {tx['action']} {tx['quantity']:.4f} {tx['ticker']} @ {tx['price_native']:.4f}"
            for tx in trades[:30]
        )
        prompt = f"""Eres el Broker Reconciliation Agent. Revisa y valida esta reconciliación XTB.

TRANSACCIONES DEL ARCHIVO ({len(trades)} total):
{trade_lines}

POSICIONES RESULTANTES:
{ticker_lines}

DEPÓSITOS EN EL PERÍODO: USD {deposits_usd:.2f}

Responde en español, máximo 60 palabras. Indica si hay anomalías (precio raro, shares negativos, ticker desconocido) o confirma que todo está correcto."""

        return _call_groq(prompt, max_tokens=200)
    except Exception as exc:
        log.warning("Groq validation failed: %s", exc)
        return None


# ── Response model ────────────────────────────────────────────────────────────

class ReconcileResult(BaseModel):
    imported: int
    errors: list[str]
    positions_updated: int
    positions_created: int
    positions_zeroed: int
    reconciled_tickers: list[str]
    deposits_usd: float
    agent_summary: str | None


# ── Endpoint ──────────────────────────────────────────────────────────────────

@router.post("/broker-reconcile", response_model=ReconcileResult)
async def broker_reconcile(
    file: UploadFile = File(...),
    user_id: str = Depends(get_user_id),
):
    """
    Upload XTB Cash Operations xlsx — wipe-and-reload strategy.
    Deletes all prior XTB transactions, re-inserts everything from the file,
    then FIFO-reconciles positions. One upload always produces correct state.
    """
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx (Cash Operations export de XTB)")

    content = await file.read()

    # 1. Parse xlsx
    try:
        trades, deposits_usd = _parse_xtb_xlsx(content)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {exc}")

    if not trades:
        raise HTTPException(status_code=422, detail="No se encontraron transacciones en el archivo.")

    db = get_admin_client()
    errors: list[str] = []

    # 2. Wipe all XTB-sourced transactions for this user
    try:
        db.table("transactions") \
          .delete() \
          .eq("user_id", user_id) \
          .like("comment", "XTB:%") \
          .execute()
    except Exception as exc:
        log.error("Could not wipe XTB transactions: %s", exc)
        raise HTTPException(status_code=500, detail=f"Error limpiando transacciones previas: {exc}")

    # 3. Insert all trades from the file
    imported = 0
    for trade in trades:
        try:
            db.table("transactions").insert({
                "user_id":      user_id,
                "ticker":       trade["ticker"],
                "action":       trade["action"],
                "quantity":     trade["quantity"],
                "price_native": trade["price_native"],
                "fee_native":   0.0,
                "currency":     _ticker_exchange_info(trade["ticker"])["currency"],
                "date":         trade["date"],
                "comment":      f"XTB: {trade['comment_raw'][:120]}",
            }).execute()
            imported += 1
        except Exception as exc:
            errors.append(f"{trade['ticker']} {trade['date']}: {exc}")

    # 4. FIFO-reconcile positions from ALL transactions (XTB + any manual ones)
    reconciled = _reconcile_positions(user_id, db)

    # 5. Upsert positions
    pos_res = db.table("positions").select("*").eq("user_id", user_id).execute()
    existing_positions: dict[str, dict] = {p["ticker"]: p for p in (pos_res.data or [])}

    positions_updated = 0
    positions_created = 0
    positions_zeroed  = 0

    # Zero out any ticker that now has 0 shares (fully sold)
    for ticker, computed in reconciled.items():
        if computed["shares"] > 0:
            continue
        for alias in _ticker_aliases(ticker):
            if alias in existing_positions:
                if float(existing_positions[alias].get("shares", 0)) > 0:
                    db.table("positions") \
                      .update({"shares": 0, "avg_cost_native": None}) \
                      .eq("user_id", user_id) \
                      .eq("ticker", alias) \
                      .execute()
                    positions_zeroed += 1
                break

    # Upsert tickers with shares > 0
    for ticker, computed in reconciled.items():
        if computed["shares"] <= 0:
            continue

        matched_key = next(
            (a for a in _ticker_aliases(ticker) if a in existing_positions),
            None,
        )

        upsert_data = {
            "shares":          computed["shares"],
            "avg_cost_native": computed["avg_cost_native"] or None,
        }

        if matched_key is not None:
            db.table("positions") \
              .update(upsert_data) \
              .eq("user_id", user_id) \
              .eq("ticker", matched_key) \
              .execute()
            positions_updated += 1
        else:
            db.table("positions").insert({
                "user_id":  user_id,
                "ticker":   ticker,
                "currency": computed["currency"],
                "market":   computed["market"],
                **upsert_data,
            }).execute()
            positions_created += 1

    # 6. Enrich names for all active positions (no-op if name already set)
    active_tickers = [t for t, v in reconciled.items() if v["shares"] > 0]
    _enrich_names(active_tickers, db, user_id)

    # 7. Groq validation
    agent_summary = _groq_validate(trades, reconciled, deposits_usd)

    log.info(
        "Broker reconcile %s: %d trades imported, %d updated, %d created, %d zeroed",
        user_id[:8], imported, positions_updated, positions_created, positions_zeroed,
    )

    return ReconcileResult(
        imported=imported,
        errors=errors[:10],
        positions_updated=positions_updated,
        positions_created=positions_created,
        positions_zeroed=positions_zeroed,
        reconciled_tickers=[t for t, v in reconciled.items() if v["shares"] > 0],
        deposits_usd=deposits_usd,
        agent_summary=agent_summary,
    )
