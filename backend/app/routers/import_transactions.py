"""
POST /api/import/xtb-xlsx  — XTB Cash Operations Excel report
"""
from __future__ import annotations

import io
import logging
import re
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel

from app.auth.dependencies import get_user_id
from app.db.supabase_client import get_admin_client

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/import", tags=["import"])


# ── XTB xlsx import ────────────────────────────────────────────────────────────

def _xtb_ticker(xtb: str) -> str:
    """Convert XTB ticker format to app/yfinance format."""
    xtb = xtb.strip()
    if xtb.endswith(".US"):
        return xtb[:-3]          # VOO.US → VOO
    if xtb.endswith(".UK"):
        return xtb[:-3] + ".L"  # EIMI.UK → EIMI.L
    return xtb                   # QDVE.DE, ZPRV.DE → unchanged


def _parse_xtb_comment(comment: str) -> tuple[float, float] | None:
    """
    Extract (quantity, price) from XTB comment field.
    Handles:
      "OPEN BUY 0.5671 @ 590.58"
      "OPEN BUY 3/3.3674 @ 30.06"   ← partial fill: quantity = 3
      "CLOSE BUY 0.0897/1.6866 @ 93.6600"
    """
    m = re.search(
        r'(?:OPEN|CLOSE)\s+(?:BUY|SELL)\s+([\d.]+)(?:/[\d.]+)?\s*@\s*([\d.]+)',
        comment or "",
    )
    if not m:
        return None
    return float(m.group(1)), float(m.group(2))


class XTBImportResult(BaseModel):
    imported: int
    skipped: int
    errors: list[str]
    tickers: list[str]
    deposits_usd: float


@router.post("/xtb-xlsx", response_model=XTBImportResult)
async def import_xtb_xlsx(
    file: UploadFile = File(...),
    user_id: str = Depends(get_user_id),
):
    """
    Upload an XTB 'Cash Operations' Excel report (.xlsx) to import trades
    as BUY/SELL transactions. Deposits are counted for reference only.
    """
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be a .xlsx XTB Cash Operations export")

    content_bytes = await file.read()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read xlsx: {exc}")

    # XTB exports to 'Cash Operations' sheet
    ws = wb.active

    db = get_admin_client()
    imported = 0
    skipped = 0
    errors: list[str] = []
    tickers: set[str] = set()
    deposits_usd = 0.0

    # Find the header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "Type":
            header_row = i
            break

    if header_row is None:
        raise HTTPException(status_code=422, detail="Could not find header row. Expected 'Type' column.")

    # Parse data rows
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or row[0] is None:
            continue
        tx_type_raw = str(row[0]).strip()
        ticker_raw = str(row[1] or "").strip()
        time_val = row[3]
        amount = row[4]
        comment = str(row[6] or "").strip()

        # Track deposits for capital reference
        if tx_type_raw == "Deposit":
            try:
                deposits_usd += float(amount or 0)
            except Exception:
                pass
            skipped += 1
            continue

        # Skip non-trade rows
        if tx_type_raw not in ("Stock purchase", "Stock sell"):
            skipped += 1
            continue

        if not ticker_raw:
            skipped += 1
            continue

        # Parse comment for qty/price
        parsed = _parse_xtb_comment(comment)
        if parsed is None:
            errors.append(f"Could not parse comment: {comment[:60]}")
            skipped += 1
            continue
        qty, price = parsed

        # Parse date
        if isinstance(time_val, datetime):
            tx_date = time_val.date().isoformat()
        else:
            try:
                tx_date = str(time_val)[:10]
            except Exception:
                skipped += 1
                continue

        # Convert ticker
        app_ticker = _xtb_ticker(ticker_raw)
        action = "BUY" if tx_type_raw == "Stock purchase" else "SELL"

        # Currency: XTB reports amounts in account currency (USD for this account)
        # Determine native currency from ticker exchange
        if ticker_raw.endswith(".DE"):
            native_ccy = "EUR"
        elif ticker_raw.endswith(".UK"):
            native_ccy = "GBP"
        else:
            native_ccy = "USD"

        try:
            tx_row = {
                "user_id": user_id,
                "ticker": app_ticker,
                "action": action,
                "quantity": qty,
                "price_native": price,
                "currency": native_ccy,
                "fee_native": 0.0,
                "date": tx_date,
                "comment": f"XTB: {comment[:80]}",
            }
            db.table("transactions").insert(tx_row).execute()
            tickers.add(app_ticker)
            imported += 1
        except Exception as exc:
            errors.append(f"{app_ticker} {tx_date}: {exc}")
            skipped += 1

    log.info(
        "XTB import for %s: %d imported, %d skipped, deposits=%.2f",
        user_id[:8], imported, skipped, deposits_usd,
    )

    return XTBImportResult(
        imported=imported,
        skipped=skipped,
        errors=errors[:10],
        tickers=sorted(tickers),
        deposits_usd=deposits_usd,
    )
